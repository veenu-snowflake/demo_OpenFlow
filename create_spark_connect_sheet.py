import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Snowpark Connect Q&A"

header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
category_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
category_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
data_font = Font(name="Calibri", size=10)
pitch_font = Font(name="Calibri", size=10, italic=True, color="1F4E79")
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

headers = ["#", "Category", "Question", "Answer (Layman Language)", "Pitch Line"]
col_widths = [6, 22, 45, 65, 55]

for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = "A1:E1"

data = [
    # 1. BASICS
    ("1. BASICS / WHAT IS IT", None),
    ("1.1", "Basics", "What is Snowpark Connect?", "It lets you run your existing PySpark code directly on Snowflake's engine — no Spark cluster needed. You change the connection string, not the code.", "Your Spark code, Snowflake's engine — zero rewrite, zero clusters."),
    ("1.2", "Basics", "Is it actually running Spark under the hood?", "No. Your PySpark client sends a logical plan to Snowflake, and Snowflake's vectorized engine executes it. There's no Spark cluster anywhere.", "Snowflake speaks Spark so you don't have to speak Snowflake."),
    ("1.3", "Basics", "When did it go GA?", "Public Preview in Jul 2025. GA on Nov 4, 2025.", "Production-ready since November 2025."),
    ("1.4", "Basics", "What Spark versions does it support?", "Spark 3.5.x only.", "Compatible with the latest Spark 3.5 APIs."),
    ("1.5", "Basics", "What languages are supported?", "PySpark only today. Java/Scala support is under development.", "PySpark today, Java/Scala on the roadmap."),

    # 2. ARCHITECTURE
    ("2. ARCHITECTURE / HOW IT WORKS", None),
    ("2.1", "Architecture", "Where does the compute happen?", "On Snowflake virtual warehouses — the same ones that run your SQL queries. No separate Spark cluster.", "One warehouse, all workloads — SQL and Spark, same engine."),
    ("2.2", "Architecture", "Do I need to manage min/max nodes?", "No. You just pick a warehouse size (XS to 4XL). Snowflake handles scaling, tuning, and auto-suspend.", "Size a warehouse, not a cluster."),
    ("2.3", "Architecture", "How does the Spark Connect protocol work?", "Spark 3.4 introduced a client-server split. Your PySpark app is the client. It sends an unresolved logical plan over gRPC. Snowflake acts as the server and executes it.", "Your app talks Spark protocol; Snowflake does the heavy lifting."),
    ("2.4", "Architecture", "Is there data movement?", "No. Data stays in Snowflake. Unlike the old Spark Connector, nothing gets extracted to an external cluster.", "Process data where it lives — no extract, no egress, no lag."),
    ("2.5", "Architecture", "What development tools can I use?", "Jupyter Notebooks, VS Code, Apache Airflow, Snowflake Notebooks, and Snowpark Submit (CLI for batch jobs).", "Use your IDE, your orchestrator, your notebook — just connect to Snowflake."),

    # 3. MIGRATION
    ("3. MIGRATION / WHY MOVE", None),
    ("3.1", "Migration", "Why not just use Snowpark instead?", "Snowpark requires you to learn a new API and rewrite code. Snowpark Connect lets you keep your existing PySpark code as-is.", "Snowpark Connect = zero-rewrite migration. Snowpark = new projects."),
    ("3.2", "Migration", "What if I have 100K lines of PySpark on Databricks?", "Point it at Snowflake instead of your Spark cluster. Most DataFrame and SQL code works without changes.", "Change the connection, keep the code — migrate in days, not quarters."),
    ("3.3", "Migration", "How much faster is it vs managed Spark?", "Snowflake claims 5.6x faster and 41% cheaper based on real customer production workloads (Nov 2022 – May 2025).", "5.6x faster, 41% cheaper — validated on real customer pipelines."),
    ("3.4", "Migration", "What about governance during migration?", "Since everything runs on Snowflake, you get RBAC, masking, row access policies, and audit trails on day one — no separate governance layer.", "Migrate your Spark code, inherit enterprise governance for free."),
    ("3.5", "Migration", "Can I run Snowpark and Snowpark Connect together?", "Yes. Same warehouse, same data. Teams can use whichever API they prefer.", "No forklift — adopt incrementally, team by team."),

    # 4. SUPPORTED FEATURES
    ("4. SUPPORTED FEATURES", None),
    ("4.1", "Features", "What Spark APIs are supported?", "DataFrame API, Spark SQL, and UDFs (Python). Covers the vast majority of real-world analytical workloads.", "DataFrames, SQL, UDFs — the APIs that matter, ready today."),
    ("4.2", "Features", "Does it support Iceberg tables?", "Yes — Snowflake-managed Iceberg, externally-managed Iceberg, and catalog-linked databases all work.", "Open Iceberg tables, Snowflake performance — via Spark syntax."),
    ("4.3", "Features", "Can I read/write CSV, JSON, Parquet?", "Yes. CSV, JSON, Parquet, Text, and XML are supported for both read and write (with some option limitations).", "Standard file formats, standard Spark code — just works."),
    ("4.4", "Features", "Are Python UDFs supported?", "Yes. You can write and register Python UDFs, including importing external libraries from Snowflake stages.", "Bring your Python logic — UDFs run natively on Snowflake."),
    ("4.5", "Features", "Does it support Spark SQL?", "Yes, full Spark SQL support including CREATE TABLE, SELECT, JOINs, aggregations, window functions, etc.", "Write Spark SQL, run on Snowflake — no translation needed."),

    # 5. LIMITATIONS
    ("5. LIMITATIONS / WHAT'S NOT SUPPORTED", None),
    ("5.1", "Limitations", "Does it support RDDs?", "No. RDDs are not supported (they weren't supported in Spark Connect itself either).", "RDDs are legacy. DataFrames are the future — and they're fully supported."),
    ("5.2", "Limitations", "What about Spark Structured Streaming?", "Not supported. For real-time streaming, use Snowpipe Streaming or OpenFlow instead.", "Streaming belongs in OpenFlow. Processing belongs in Snowpark Connect."),
    ("5.3", "Limitations", "Can I use MLlib / Spark ML?", "No. For ML, use Snowpark ML, Snowflake Model Registry, or Cortex AI instead.", "Skip MLlib. Snowpark ML trains, registers, and deploys — all in Snowflake."),
    ("5.4", "Limitations", "What about Delta Lake APIs?", "Not supported. Snowflake uses Iceberg as the open table format instead.", "Iceberg is the open standard. Snowflake bets on open, not proprietary."),
    ("5.5", "Limitations", "Can I tune executor memory, shuffle partitions?", "No. Snowflake's optimizer handles all of that automatically. Hints and repartition calls are silently ignored (no-op).", "No tuning knobs to turn. Snowflake optimizes for you — that's the point."),
    ("5.6", "Limitations", "Java/Scala Spark code?", "Not yet. PySpark only today. Java/Scala is on the roadmap.", "PySpark first. Java/Scala next. Most Spark shops are Python-first anyway."),
    ("5.7", "Limitations", "Avro and ORC file formats?", "Not supported. Use Parquet, CSV, JSON, Text, or XML instead.", "Parquet is the industry standard — and it's fully supported."),

    # 6. COST & PRICING
    ("6. COST & PRICING", None),
    ("6.1", "Cost", "How is it priced?", "Same as any Snowflake warehouse — credit-based, per-second billing. No separate Spark licensing or cluster cost.", "One bill. Per-second. No Spark license, no cluster overhead."),
    ("6.2", "Cost", "Is there additional cost for Snowpark Connect?", "No extra fee for the feature itself. You pay for warehouse compute time, same as SQL.", "No premium — Spark workloads priced exactly like SQL workloads."),
    ("6.3", "Cost", "How does it compare to Databricks pricing?", "No cluster management cost, no idle cluster waste, no DBU markups. Warehouse auto-suspends when idle.", "Pay when you compute, not when you wait. No idle cluster burns."),
    ("6.4", "Cost", "Can I use auto-suspend/auto-resume?", "Yes. Warehouse spins down when your Spark job finishes — no manual cluster teardown.", "Job done? Warehouse sleeps. That's money back in your pocket."),

    # 7. VS OTHER OFFERINGS
    ("7. VS OTHER SNOWFLAKE OFFERINGS", None),
    ("7.1", "Comparison", "Snowpark Connect vs Snowpark?", "Snowpark = Snowflake's native API (learn new syntax). Snowpark Connect = run existing PySpark as-is. Same engine underneath.", "New projects -> Snowpark. Existing Spark code -> Snowpark Connect."),
    ("7.2", "Comparison", "Snowpark Connect vs Spark Connector?", "Spark Connector moves data OUT to your cluster. Snowpark Connect keeps data IN Snowflake and runs it there.", "Connector = data moves to you. Connect = you move to the data."),
    ("7.3", "Comparison", "Snowpark Connect vs SPCS + Spark?", "SPCS = run a real DIY Spark cluster in containers (you manage everything). Snowpark Connect = serverless, Snowflake manages everything.", "SPCS is a container. Snowpark Connect is a service."),
    ("7.4", "Comparison", "When would I still use the old Spark Connector?", "Only if you need to keep Spark on your own infrastructure (Databricks, EMR) and just read/write Snowflake data.", "If you love your cluster, keep the Connector. If you want to drop it, use Connect."),

    # 8. SECURITY
    ("8. SECURITY & GOVERNANCE", None),
    ("8.1", "Security", "How is data secured?", "Same as any Snowflake workload — TLS in transit, AES-256 at rest, end-to-end encryption.", "Spark code inherits Snowflake security — nothing bolted on."),
    ("8.2", "Security", "Do RBAC policies apply to Spark queries?", "Yes. All role-based access, masking policies, and row access policies apply to Spark queries just like SQL.", "One governance framework. SQL or Spark — same rules, same audit trail."),
    ("8.3", "Security", "Does it work with PrivateLink?", "Yes. Snowpark Connect runs through the same Snowflake endpoint — PrivateLink, IP whitelisting, all apply.", "Private network, private data, public Spark APIs."),
    ("8.4", "Security", "What about compliance (SOC2, HIPAA)?", "Since it runs on Snowflake infrastructure, it inherits all Snowflake compliance certifications.", "SOC2, HIPAA, FedRAMP — inherited, not re-certified."),

    # 9. USE CASES
    ("9. REAL-WORLD USE CASES", None),
    ("9.1", "Use Cases", "Migrating off Databricks — is this the path?", "Yes. Snowpark Connect is explicitly designed for this. Move PySpark workloads to Snowflake without rewriting code.", "Databricks exit strategy: Snowpark Connect. Days, not months."),
    ("9.2", "Use Cases", "Migrating off EMR?", "Same story. Point your PySpark jobs at Snowflake instead of EMR. Eliminate cluster ops entirely.", "No more EMR cluster babysitting. Submit your Spark job, go home."),
    ("9.3", "Use Cases", "Can I use it for batch ETL?", "Yes. Use Snowpark Submit to run PySpark batch jobs asynchronously on Snowflake warehouses, orchestrated by Airflow.", "Batch ETL without batch infrastructure."),
    ("9.4", "Use Cases", "Interactive data exploration?", "Yes. Use Jupyter or Snowflake Notebooks with PySpark syntax for interactive analysis.", "Explore with PySpark in a notebook, powered by Snowflake speed."),
    ("9.5", "Use Cases", "Can I use it with Airflow?", "Yes. Airflow can trigger Snowpark Connect jobs just like it triggers any Spark job — via Snowpark Submit CLI.", "Airflow orchestrates, Snowflake executes — keep your DAGs, drop your clusters."),

    # 10. OBJECTION HANDLERS
    ("10. CUSTOMER OBJECTION HANDLERS", None),
    ("10.1", "Objections", "\"We're happy on Databricks\"", "Are you happy with the bill? Snowpark Connect is 41% cheaper with 5.6x better performance on real workloads. And you keep your PySpark code.", "Keep the code. Drop the cost. That's the conversation."),
    ("10.2", "Objections", "\"What about Streaming?\"", "Spark Streaming isn't supported, but Snowflake has Snowpipe Streaming and OpenFlow for real-time ingestion. Separate concern.", "Ingest real-time with OpenFlow. Process with Snowpark Connect. Best of both."),
    ("10.3", "Objections", "\"We need MLlib\"", "Snowpark ML + Model Registry replaces MLlib with a managed experience. Train, register, deploy — all in Snowflake.", "MLlib is DIY ML. Snowpark ML is managed ML. Same Snowflake bill."),
    ("10.4", "Objections", "\"It's only PySpark, we have Scala\"", "80%+ of Spark workloads today are PySpark. Java/Scala is on the roadmap. Start with PySpark, migrate the rest later.", "Start with PySpark — that's where most of your Spark code lives anyway."),
    ("10.5", "Objections", "\"What if we hit a compatibility issue?\"", "Snowflake publishes a detailed compatibility guide. Most gaps are edge cases (NullType, Iterator UDFs, Avro/ORC). Core DataFrame + SQL works.", "99% of real-world Spark code works. The 1% has documented workarounds."),
    ("10.6", "Objections", "\"We don't want vendor lock-in\"", "Your code is standard PySpark. It runs on any Spark Connect server. Plus Snowflake supports Iceberg — your data stays open too.", "Standard PySpark code + open Iceberg tables = zero lock-in."),
]

row = 2
for item in data:
    if item[1] is None:
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = category_fill
            cell.font = category_font
            cell.border = thin_border
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value=item[0])
        ws.cell(row=row, column=1).font = category_font
        ws.cell(row=row, column=1).fill = category_fill
        ws.cell(row=row, column=1).alignment = Alignment(vertical="center")
        ws.cell(row=row, column=1).border = thin_border
        row += 1
    else:
        num, cat, q, a, p = item
        for col, val in enumerate([num, cat, q, a, p], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = pitch_font if col == 5 else data_font
            cell.alignment = wrap
            cell.border = thin_border
        ws.row_dimensions[row].height = 45
        row += 1

out = "/Users/vyadav/Desktop/Projects/Demo/demo_OpenFlow/Snowpark-Connect-QnA-Sheet.xlsx"
wb.save(out)
print(f"Saved to {out}")
